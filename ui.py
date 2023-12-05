import tkinter
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import tkinter.font as font
import pandas as pd
import numpy as np
import sys, os, cv2
from pathlib import Path
from typing import Sequence
from skimage.metrics import structural_similarity as ssim


def detect_intro_outro_single(video_path: str | os.PathLike, flag: bool, filepath: str, genre: str) -> tuple[int, int, int, str] | None:
    """
    Your current approach of detecting intro and outro times based on the mean pixel value of grayscale frames has a few limitations. It assumes that the intro is darker (mean pixel value < 10) and the outro is brighter (mean pixel value > 100). This might not always be the case for all videos.
    Here are a few suggestions to improve the success rate:
    Use more features: Instead of just using the mean pixel value, you can use more features of the video frames. For example, you can use the standard deviation of pixel values, the color histogram, etc. This will give you a more comprehensive understanding of the video content.
    Use machine learning: If you have a labeled dataset of videos with known intro and outro times, you can train a machine learning model to predict the intro and outro times based on the features of the video frames. This will likely give you a higher success rate than a heuristic approach.
    Use audio features: Often, the intro and outro of a video have distinctive audio characteristics, such as theme music. You can extract audio features from the video and use them to detect the intro and outro times.
    Use scene change detection: Intros and outros often coincide with scene changes. You can use scene change detection algorithms to detect the intro and outro times.
    Improve the granularity of your search: Instead of breaking after the first 3 minutes, you can continue scanning the video but with a less frequent sampling rate. This might help to catch intros and outros that are longer than usual.
    """
    filename = os.path.splitext(os.path.basename(video_path))[0]
    video_path = os.fspath(video_path)
    video_capture = cv2.VideoCapture(video_path)

    threshold_upper_limit = 220

    if not video_capture.isOpened():
        return

    frame_count = int(video_capture.get(cv2.CAP_PROP_FRAME_COUNT))
    frame_width = int(video_capture.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(video_capture.get(cv2.CAP_PROP_FRAME_HEIGHT))
    video_fps = video_capture.get(cv2.CAP_PROP_FPS)
    video_duration = frame_count / video_fps

    predicted_intro_duration, predicted_outro_duration = predict_duration(video_duration, genre=genre)

    if frame_width >= 1280 and frame_height >= 710:
        video_definition = '高清'
    else:
        video_definition = '标清'

    intro_end_time = None
    outro_start_time = None
    intro_frame = None
    outro_frame = None
    frame: np.ndarray

    for frame_index in range(50, frame_count):
        ret, frame = video_capture.read()

        if not ret:
            break

        if frame_index / video_fps > predicted_intro_duration:  # Restrict to first 1 minutes
            break

        # calculates the mean value of an image or a specific region of interest (ROI)
        # src: This is the source image or the image from which you want to calculate the mean value.
        # mask (optional): This is an optional mask image that specifies the region of interest. If provided, the mean value will be calculated only for the pixels within the mask. If not provided, the mean value will be calculated for the entire image.
        # The cv2.mean function returns a tuple containing the mean values for each channel (B, G, R, and optionally alpha) of the image or ROI.
        grayscale_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        grayscale_mean = cv2.mean(grayscale_frame)
        grayscale_mean_pixel_value = grayscale_mean[0]

        mean = cv2.mean(frame)
        mean_pixel_value_b = mean[0]  # Blue channel
        mean_pixel_value_g = mean[1]  # Green channel
        mean_pixel_value_r = mean[2]  # Red channel

        if (grayscale_mean_pixel_value < 20
                # grayscale_mean_pixel_value > threshold_upper_limit or
                # mean_pixel_value_b > threshold_upper_limit or  # Check if blue channel exceeds threshold_upper_limit
                # mean_pixel_value_g > threshold_upper_limit or  # Check if green channel exceeds threshold_upper_limit
                # mean_pixel_value_r > threshold_upper_limit  # Check if red channel exceeds threshold_upper_limit
        ) and intro_end_time is None:
            print('intro', frame, grayscale_mean, mean)
            # intro_end_time = frame_index / video_fps
            intro_end_time = video_capture.get(cv2.CAP_PROP_POS_MSEC) / 1000
            intro_frame = frame.copy()
            break

    if video_duration >= 120:  # duration >= 2m
        outro_frame_start_index = int((video_duration - 60) * video_fps)
    else:
        outro_frame_start_index = frame_count // 2

    video_capture.set(cv2.CAP_PROP_POS_FRAMES, outro_frame_start_index)

    while True:
        ret, frame = video_capture.read()
        outro_frame_start_index += 1
        if not ret:
            break

        grayscale_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        grayscale_mean = cv2.mean(grayscale_frame)
        grayscale_mean_pixel_value = grayscale_mean[0]

        mean = cv2.mean(frame)
        mean_pixel_value_b = mean[0]  # Blue channel
        mean_pixel_value_g = mean[1]  # Green channel
        mean_pixel_value_r = mean[2]  # Red channel

        if (grayscale_mean_pixel_value < 20
                # grayscale_mean_pixel_value > threshold_upper_limit or
                # mean_pixel_value_b > threshold_upper_limit or  # Check if blue channel exceeds threshold_upper_limit
                # mean_pixel_value_g > threshold_upper_limit or  # Check if green channel exceeds threshold_upper_limit
                # mean_pixel_value_r > threshold_upper_limit  # Check if red channel exceeds threshold_upper_limit
        ) and outro_start_time is None:
            print('outro', frame, grayscale_mean, mean)
            # outro_start_time = (outro_frame_start_index + 1) / video_fps
            outro_start_time = video_capture.get(cv2.CAP_PROP_POS_MSEC) / 1000
            outro_frame = frame.copy()
            break

    # for frame_index in range(frame_count - 1, -1, -1):
    #     video_capture.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
    #     ret, frame = video_capture.read()
    #
    #     if not ret:
    #         break
    #
    #     if video_duration - (frame_index / video_capture.get(cv2.CAP_PROP_FPS)) > 180:  # Restrict to last 3 minutes
    #         break
    #
    #     grayscale_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    #     mean = cv2.mean(grayscale_frame)
    #     mean_pixel_value = mean[0]
    #     if mean_pixel_value > 100 and outro_start_time is None:
    #         outro_start_time = (frame_index + 1) / video_capture.get(cv2.CAP_PROP_FPS)
    #         break

    # Additional logic
    if flag:
        if intro_frame is not None:
            cv2.imwrite(os.path.join(filepath, f'{filename}-intro.jpeg'), intro_frame)
        if outro_frame is not None:
            cv2.imwrite(os.path.join(filepath, f'{filename}-outro.jpeg'), outro_frame)

    if intro_end_time is None:
        intro_end_time = 0.0
    if outro_start_time is None:
        outro_start_time = video_duration

    video_capture.release()

    return round(intro_end_time), round(outro_start_time), round(video_duration), video_definition


def is_video(filepath):
    try:
        cap = cv2.VideoCapture(filepath)
        if cap.isOpened():
            return True
        return False
    except cv2.error:
        return False


def convert_seconds_to_hms(total_seconds: int):
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    if hours > 0:
        return f'{hours:02d}:{minutes:02d}:{seconds:02d}'
    else:
        return f'{minutes:02d}:{seconds:02d}'


def extract_common_part(str1: str, str2: str) -> str:
    common = ''
    for c1, c2 in zip(str1, str2, strict=False):
        if c1 == c2:
            common += 'c1'
        else:
            break
    return common


def compare_frames(frame1, frame2):
    # convert frames to grayscale
    gray1 = cv2.cvtColor(frame1, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.cvtColor(frame2, cv2.COLOR_BGR2GRAY)
    return ssim(gray1, gray2)


config = {
    '少儿': {
        'intro_correlation_coefficient': 0.56,
        'outro_correlation_coefficient': 0.68,
        'intro_duration_mean': 26,
        'outro_duration_mean': 17.8,
        'total_duration_mean': 438,
        'total_duration_std': 359.8
    }
}


def predict_duration(duration: float, coefficient=1.7, genre: str = '少儿'):
    v = config[genre]
    intro_duration = v['intro_duration_mean'] + v['intro_correlation_coefficient'] * (duration - v['total_duration_mean']) / v['total_duration_std']
    outro_duration = v['outro_duration_mean'] + v['outro_correlation_coefficient'] * (duration - v['total_duration_mean']) / v['total_duration_std']
    return min(intro_duration * coefficient, duration * 0.5), min(outro_duration * coefficient, duration * 0.5)


def supplement(video_path: str | os.PathLike) -> tuple[int, str]:
    cap = cv2.VideoCapture(os.fspath(video_path))
    if int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)) >= 1280 and int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) >= 720:
        video_definition = '高清'
    else:
        video_definition = '标清'
    duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / cap.get(cv2.CAP_PROP_FPS)
    return round(duration), video_definition


def detect_intro_outro_double(video_paths: Sequence[str | os.PathLike], threshold=0.95, frame_output_flag: bool = False, frame_output_path: str = '') -> tuple[int, int] | None:
    filename = extract_common_part(*(os.fspath(v) for v in video_paths))
    frame_output_path = os.path.dirname(video_paths[0])

    cap1 = cv2.VideoCapture(os.fspath(video_paths[0]))
    cap2 = cv2.VideoCapture(os.fspath(video_paths[1]))

    if not cap1.isOpened() or not cap2.isOpened():
        return

    frame_count1 = int(cap1.get(cv2.CAP_PROP_FRAME_COUNT))
    frame_count2 = int(cap2.get(cv2.CAP_PROP_FRAME_COUNT))
    min_frame_count = min(frame_count1, frame_count2)
    video_fps1 = cap1.get(cv2.CAP_PROP_FPS)
    video_fps2 = cap2.get(cv2.CAP_PROP_FPS)
    video_duration1: float = frame_count1 / video_fps1
    video_duration2: float = frame_count2 / video_fps2

    # if int(cap1.get(cv2.CAP_PROP_FRAME_WIDTH)) >= 1280 and int(cap1.get(cv2.CAP_PROP_FRAME_HEIGHT)) >= 720:
    #     video_definition = '高清'
    # else:
    #     video_definition = '标清'

    # while True:
    #     ret1, frame1 = cap1.read()
    #     ret2, frame2 = cap2.read()
    #
    #     if not ret1 or not ret2:
    #         break
    #
    #     structural_similarity = compare_frames(frame1, frame2)
    #     if structural_similarity < threshold:
    #         print("Intro ends at frame:", frame_num)
    #         break
    #     # print(frame_num)
    #     frame_num += 1

    threshold_seconds = 180
    prev_frame1: np.ndarray | None = None
    prev_frame2: np.ndarray | None = None
    next_frame1: np.ndarray | None = None
    next_frame2: np.ndarray | None = None
    intro_duration = None
    outro_duration = None
    intro_duration_threshold = int(min(min_frame_count / 2, int(video_fps1 * threshold_seconds)))
    outro_duration_threshold1 = frame_count1 - int(min(frame_count1 / 2, int(video_fps1 * threshold_seconds)))
    outro_duration_threshold2 = frame_count2 - int(min(frame_count2 / 2, int(video_fps2 * threshold_seconds)))

    for frame_index in range(intro_duration_threshold):
        ret1, frame1 = cap1.read()
        ret2, frame2 = cap2.read()

        if not ret1 or not ret2:
            break

        structural_similarity = compare_frames(frame1, frame2)

        if structural_similarity < threshold:
            print(f'intro>> {structural_similarity=:.5f}, frame_index: {frame_index:<20} CAP_PROP_POS_FRAMES: {cap1.get(cv2.CAP_PROP_POS_FRAMES), cap2.get(cv2.CAP_PROP_POS_FRAMES)}')
            if frame_output_flag:
                cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[0])}-intro-{frame_index}.jpeg'), frame1)
                cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[1])}-intro-{frame_index}.jpeg'), frame2)
                if prev_frame1 is not None:
                    cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[0])}-intro-prev-{frame_index - 1}.jpeg'), prev_frame1)
                if prev_frame2 is not None:
                    cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[1])}-intro-prev-{frame_index - 1}.jpeg'), prev_frame2)
            intro_duration = cap1.get(cv2.CAP_PROP_POS_MSEC) / 1000
            break

        prev_frame1 = frame1
        prev_frame2 = frame2

    if intro_duration is None:
        intro_duration = 0

    frame_index1 = frame_count1
    frame_index2 = frame_count2

    while True:
        frame_index1 -= 1
        frame_index2 -= 1

        if frame_index1 < outro_duration_threshold1 or frame_index2 < outro_duration_threshold2:
            break

        cap1.set(cv2.CAP_PROP_POS_FRAMES, frame_index1)
        cap2.set(cv2.CAP_PROP_POS_FRAMES, frame_index2)

        ret1, frame1 = cap1.read()
        ret2, frame2 = cap2.read()

        if not ret1 or not ret2:
            continue

        structural_similarity = compare_frames(frame1, frame2)

        if structural_similarity < threshold:
            print(f'outro>> {structural_similarity=:.5f}, frame_index: {str((frame_index1, frame_index2)):<20} CAP_PROP_POS_FRAMES {cap1.get(cv2.CAP_PROP_POS_FRAMES), cap2.get(cv2.CAP_PROP_POS_FRAMES)}')
            if frame_output_flag:
                cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[0])}-outro-{frame_index1}.jpeg'), frame1)
                cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[1])}-outro-{frame_index2}.jpeg'), frame2)
                if next_frame1 is not None:
                    cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[0])}-outro-next-{frame_index1 + 1}.jpeg'), next_frame1)
                if next_frame2 is not None:
                    cv2.imwrite(os.path.join(frame_output_path, f'{os.path.basename(video_paths[1])}-outro-next-{frame_index2 + 1}.jpeg'), next_frame2)
            outro_duration = video_duration1 - cap1.get(cv2.CAP_PROP_POS_MSEC) / 1000
            # print(f'{video_duration1 - cap1.get(cv2.CAP_PROP_POS_MSEC) / 1000}, {video_duration2 - cap2.get(cv2.CAP_PROP_POS_MSEC) / 1000}')
            break

        next_frame1 = frame1
        next_frame2 = frame2

    if outro_duration is None:
        outro_duration = 0

    cap1.release()
    cap2.release()

    return intro_duration, outro_duration


class Application(Tk):
    def __init__(self):
        super().__init__()
        self.combobox: ttk.Combobox | None = None
        self.title('video-analyzer')
        self.geometry('900x300')
        # self.configure(background='silver')
        self.font = font.Font(name='Courier', family='Courier', size=20, weight='normal')
        self.option_add('*Font', '"{family}" {size} {weight} {slant}'.format(**self.font.actual()))
        self.values = ['少儿', '电影', '电视剧', '纪录片', '综艺']
        self.flag_askdirectory = False
        self.flag_asksaveasfilename = False
        self.flag_output_intro_outro_frame = BooleanVar()
        self.askdirectory = StringVar(self, value='还未选择视频源', name='askdirectory')
        self.asksaveasfilename = StringVar(self, value='还未选择输出文件', name='asksaveasfilename')
        self.init()
        self.init_widget()
        self.init_style()

    def init(self):
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        # width = int(screen_width * .4)
        # height = int(screen_height * .4)
        width = 1300
        height = 350
        offset_top = int(screen_height / 2 - height / 2)
        offset_left = int(screen_width / 2 - width / 2)
        self.geometry(f'{width}x{height}+{offset_left}+{offset_top}')

    def init_widget(self):
        ttk.Button(self, text='请选择视频源', command=self.cb_askdirectory, style='TButton').grid(row=0, column=0, ipadx=2, ipady=2, padx=8, pady=10)
        ttk.Label(self, textvariable=self.askdirectory).grid(row=0, column=1, sticky='w')
        ttk.Button(self, text='请选择输出文件', command=self.cb_asksaveasfile, style='TButton').grid(row=1, column=0, ipadx=2, ipady=2, padx=8, pady=10)
        ttk.Label(self, textvariable=self.asksaveasfilename).grid(row=1, column=1, sticky='w')
        ttk.Checkbutton(self, text='是否输出片头和片尾画面', variable=self.flag_output_intro_outro_frame).grid(row=2, column=0, padx=8, pady=10, sticky='w')
        frame = ttk.Frame(self)
        frame.grid(row=2, column=1, padx=0, pady=10, sticky='w')
        # ttk.Label(frame, text='请选择类型').grid(row=2, column=1, padx=8, pady=10, sticky='w')
        ttk.Label(frame, text='请选择类型').pack(side='left')
        combobox = ttk.Combobox(frame, values=self.values, textvariable=tkinter.StringVar())
        combobox.set('少儿')
        combobox.pack(side='left', padx=8)
        combobox.bind('<<ComboboxSelected>>', lambda e: print(e, combobox.get()))
        self.combobox = combobox
        ttk.Button(self, text='运行', name='run', command=self.cb_run).grid(row=3, column=0)
        ttk.Button(self, text='退出', command=self.cb_exit).grid(row=3, column=1, sticky='w')
        ttk.Label(self, text='视频分析会占用大量cpu, 会出现卡死情况, 请不要关闭软件, \n建议在空闲时间使用, 避免影响正常工作').grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky='w')
        self.grid_columnconfigure(index=0, minsize=20, weight=0)
        self.grid_columnconfigure(index=1, minsize=200, weight=1)

    def cb_askdirectory(self):
        askdirectory = filedialog.askdirectory(initialdir=Path.home() / 'Desktop', mustexist=True, title='请选择视频源')
        print(f'{askdirectory=}')
        if askdirectory:
            self.askdirectory.set(askdirectory)
            self.flag_askdirectory = True

    def cb_asksaveasfile(self):
        asksaveasfilename = filedialog.asksaveasfilename(confirmoverwrite=True, defaultextension='xlsx', initialdir=Path.home() / 'Desktop', initialfile='result.xlsx', title='asksaveasfilename')
        print(f'{asksaveasfilename=}')
        if asksaveasfilename:
            self.asksaveasfilename.set(asksaveasfilename)
            self.flag_asksaveasfilename = True

    def cb_run(self):
        frame_output_path = os.path.dirname(self.asksaveasfilename.get())
        genre = self.combobox.get()

        if genre not in self.values:
            messagebox.showwarning(title='showwarning', message='类型选择有误!')
            return
        if genre not in config:
            messagebox.showerror(title='showerror', message=f'{genre}类型逻辑暂未实现')
            return
        if not self.flag_askdirectory:
            messagebox.showwarning(title='showwarning', message='您还未选择视频源!')
            return
        if not self.flag_asksaveasfilename:
            messagebox.showwarning(title='showwarning', message='您还未选择输出文件!')
            return

        print(self.askdirectory.get(), self.asksaveasfilename.get())
        # btn_run: Widget = self.children['run']
        self.children['run'].configure({'state': 'disabled'})

        records = {'文件名称': [], '文件格式': [], '片头结束时间': [], '片尾开始时间': [], '片头时长': [], '片尾时长': [], '时长(秒)': [], '时长(分钟)': [], '运行时间': [], '清晰度': [], '文件大小(KiB)': [], '文件大小(MiB)': []}

        try:
            for dirpath, dirnames, filenames in os.walk(self.askdirectory.get(), topdown=False, followlinks=True):
                print(dirpath, dirnames, filenames)
                print('~' * 80)
                if not dirnames:
                    video_paths = []
                    for filename in filenames:
                        filename: str
                        # if re.fullmatch('[0-9_]',os.path.splitext(filename)[0]):
                        filepath = os.path.join(dirpath, filename)
                        if is_video(filepath):
                            video_paths.append(filepath)
                    if len(video_paths) >= 2:
                        intro_duration, outro_duration = detect_intro_outro_double(video_paths[:2], frame_output_flag=self.flag_output_intro_outro_frame.get(), frame_output_path=frame_output_path)
                        for video_path in video_paths:
                            filesize = os.path.getsize(video_path)
                            kilobytes = round(filesize / 1024, 2)
                            megabytes = round(kilobytes / 1024, 2)
                            duration, video_definition = supplement(video_path)
                            path = Path(video_path)
                            records['文件名称'].append(path.stem)
                            records['文件格式'].append(path.suffix.replace('.', ''))
                            records['片头结束时间'].append(intro_duration)
                            records['片尾开始时间'].append(int(duration - outro_duration))
                            records['片头时长'].append(intro_duration)
                            records['片尾时长'].append(outro_duration)
                            records['时长(秒)'].append(duration)
                            records['时长(分钟)'].append(round(duration / 60, 2))
                            records['运行时间'].append(convert_seconds_to_hms(duration))
                            records['清晰度'].append(video_definition)
                            records['文件大小(KiB)'].append(kilobytes)
                            records['文件大小(MiB)'].append(megabytes)
                    elif len(video_paths) == 1:
                        path = Path(video_paths[0])
                        v = detect_intro_outro_single(path, self.flag_output_intro_outro_frame.get(), os.path.dirname(self.asksaveasfilename.get()), self.combobox.get())
                        if v:
                            filesize = os.path.getsize(path)
                            kilobytes = round(filesize / 1024, 2)
                            megabytes = round(kilobytes / 1024, 2)
                            # record = Record(filename.stem, filename.suffix.replace('.', ''), v[0], v[1], v[0], v[2] - v[1], v[2], v[2] // 60, convert_seconds_to_hms(v[2]), v[3], kilobytes, megabytes)
                            # records.append(record)
                            records['文件名称'].append(path.stem)
                            records['文件格式'].append(path.suffix.replace('.', ''))
                            records['片头结束时间'].append(v[0])
                            records['片尾开始时间'].append(v[1])
                            records['片头时长'].append(v[0])
                            records['片尾时长'].append(v[2] - v[1])
                            records['时长(秒)'].append(v[2])
                            records['时长(分钟)'].append(round(v[2] / 60, 2))
                            records['运行时间'].append(convert_seconds_to_hms(v[2]))
                            records['清晰度'].append(v[3])
                            records['文件大小(KiB)'].append(kilobytes)
                            records['文件大小(MiB)'].append(megabytes)

            # try:
            #     for filename in Path(self.askdirectory.get()).iterdir():
            #         if filename.is_file():
            #             v = detect_intro_outro_single(filename, self.flag_output_intro_outro_frame.get(), os.path.dirname(self.asksaveasfilename.get()))
            #             if v:
            #                 filesize = os.path.getsize(filename)
            #                 kilobytes = round(filesize / 1024, 2)
            #                 megabytes = round(kilobytes / 1024, 2)
            #                 # record = Record(filename.stem, filename.suffix.replace('.', ''), v[0], v[1], v[0], v[2] - v[1], v[2], v[2] // 60, convert_seconds_to_hms(v[2]), v[3], kilobytes, megabytes)
            #                 # records.append(record)
            #                 records['文件名称'].append(filename.stem)
            #                 records['文件格式'].append(filename.suffix.replace('.', ''))
            #                 records['片头结束时间'].append(v[0])
            #                 records['片尾开始时间'].append(v[1])
            #                 records['片头时长'].append(v[0])
            #                 records['片尾时长'].append(v[2] - v[1])
            #                 records['时长(秒)'].append(v[2])
            #                 records['时长(分钟)'].append(round(v[2] / 60, 2))
            #                 records['运行时间'].append(convert_seconds_to_hms(v[2]))
            #                 records['清晰度'].append(v[3])
            #                 records['文件大小(KiB)'].append(kilobytes)
            #                 records['文件大小(MiB)'].append(megabytes)
            df = pd.DataFrame(records)
            sheet_name = os.path.basename(self.askdirectory.get())
            df.to_excel(self.asksaveasfilename.get(), sheet_name=sheet_name, index=False)
            self.beautify()
            messagebox.showinfo(title='showinfo', message='视频分析已经完成')
        except Exception as e:
            messagebox.showerror(title='showerror', message=str(e))
        finally:
            self.children['run'].configure({'state': 'normal'})

    def beautify(self):
        wb: Workbook = load_workbook(self.asksaveasfilename.get())
        ws = wb.active
        # Iterate over the columns
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length * 2 + 8)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Center align the text
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        wb.save('test.xlsx')
        wb.close()

    def cb_exit(self):
        self.quit()

    def bind_event(self):
        self.bind('<Control-q>', lambda e: e.widget.destroy())

    def init_style(self):
        style = ttk.Style()
        # style.configure('TButton', foreground='black', background='white', font=('Helvetica', 38, 'bold'), padding=5)
        style.configure('TButton', width=15, relief='flat', font=('Helvetica', 18, 'bold'), padding=5)
        style.configure('TButton.Disabled', background='gray')
        style.map('TButton', foreground=[('disabled', 'gray'), ('!disabled', 'black')])


app = Application()

app.mainloop()
